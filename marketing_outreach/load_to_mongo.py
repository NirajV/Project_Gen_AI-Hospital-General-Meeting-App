"""
Load Apollo.io / ZoomInfo / Hunter / LinkedIn-format XLSX or CSV contact data
into a MongoDB collection: `marketing_contacts`.

This database / collection is **completely separate** from the BioMedMeet
application's collections (meetings, patients, users, etc.). It lives in its
own MongoDB database called `biomedmeet_marketing` by default — set
MARKETING_DB_URL and MARKETING_DB_NAME in .env to override.

USAGE
=====
    # First import — replaces the entire collection
    python3 -m marketing_outreach.load_to_mongo \
        --input data/Hospital_Contact_Detail_V2.xlsx --replace

    # Upsert (merge) — keeps existing rows that aren't in the new file
    python3 -m marketing_outreach.load_to_mongo \
        --input data/Hospital_Contact_Detail_V2.xlsx --upsert

    # Filter on import (e.g., only US contacts with verified emails)
    python3 -m marketing_outreach.load_to_mongo \
        --input data/Hospital_Contact_Detail_V2.xlsx \
        --country US --email-status Verified --replace

After loading, send_campaign.py can read from MongoDB with --source mongo.
"""
from __future__ import annotations

import argparse
import os
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv

HERE = Path(__file__).resolve().parent
load_dotenv(HERE / ".env")

# Marketing DB lives in its own database — fully isolated from the app
MARKETING_DB_URL = os.environ.get("MARKETING_DB_URL", "mongodb://localhost:27017")
MARKETING_DB_NAME = os.environ.get("MARKETING_DB_NAME", "biomedmeet_marketing")
COLLECTION = "marketing_contacts"


# ---------------------------------------------------------------------------
# Apollo XLSX column → BioMedMeet schema mapping
# ---------------------------------------------------------------------------
# Apollo's verified-export columns are documented here. We preserve everything
# we might need later but flatten the most important fields to predictable keys.

APOLLO_FIELDS = {
    "First Name":              "first_name",
    "Last Name":               "last_name",
    "Title":                   "title",
    "Company Name":            "company",
    "Email":                   "email",
    "Email Status":            "email_status",
    "Email Confidence":        "email_confidence",
    "Seniority":               "seniority",
    "Departments":             "departments",
    "Sub Departments":         "sub_departments",
    "Work Direct Phone":       "phone_direct",
    "Mobile Phone":            "phone_mobile",
    "Corporate Phone":         "phone_corporate",
    "Person Linkedin Url":     "linkedin_url",
    "Website":                 "company_website",
    "Company Linkedin Url":    "company_linkedin",
    "City":                    "person_city",
    "State":                   "person_state",
    "Country":                 "person_country",
    "Company City":            "company_city",
    "Company State":           "company_state",
    "Company Country":         "company_country",
    "Company Address":         "company_address",
    "# Employees":             "company_employees",
    "Industry":                "industry",
    "Annual Revenue":          "company_revenue",
}


def _normalise_country(value: str | None) -> str:
    """Map free-text country to two-letter region codes used by send_campaign."""
    if not value:
        return ""
    v = value.strip().lower()
    if v in {"united states", "usa", "us", "united states of america"}:
        return "US"
    if v in {"united kingdom", "uk", "great britain", "england", "scotland", "wales"}:
        return "UK"
    if v in {"india", "in"}:
        return "IN"
    if v in {"germany", "france", "spain", "italy", "netherlands", "sweden",
             "switzerland", "denmark", "norway", "finland", "ireland", "belgium",
             "portugal", "austria", "poland"}:
        return "EU"
    return value.strip().upper()[:2]


def _row_to_doc(row: dict) -> dict:
    """Convert an Apollo XLSX row dict to our normalised contact document."""
    doc = {}
    for src, dst in APOLLO_FIELDS.items():
        v = row.get(src)
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        doc[dst] = v.strip() if isinstance(v, str) else v

    # Synthesised fields used by the email template
    first = doc.get("first_name", "").strip()
    last = doc.get("last_name", "").strip()
    doc["contact_first_name"] = first
    doc["contact_person"] = f"{first} {last}".strip()
    doc["digital_team_name"] = doc.get("title", "") or doc.get("departments", "")
    doc["hospital_name"] = doc.get("company", "")
    doc["contact_email"] = (doc.get("email", "") or "").lower()
    doc["phone"] = (
        doc.get("phone_direct")
        or doc.get("phone_mobile")
        or doc.get("phone_corporate")
        or ""
    )
    if isinstance(doc["phone"], str):
        doc["phone"] = doc["phone"].lstrip("'")  # Apollo prefixes phone with '
    doc["country"] = _normalise_country(doc.get("person_country") or doc.get("company_country"))

    # Send-status metadata used by send_campaign.py
    doc["status"] = "pending"           # pending | sent | failed | unsubscribed | bounced
    doc["last_attempt_at"] = None
    doc["last_error"] = None
    doc["imported_at"] = datetime.now(timezone.utc)

    return doc


def load_xlsx(path: Path) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h is not None else "" for h in rows[0]]
    docs = []
    for raw in rows[1:]:
        row = {headers[i]: raw[i] for i in range(min(len(headers), len(raw)))}
        if not row.get("Email"):
            continue
        docs.append(_row_to_doc(row))
    return docs


def load_csv(path: Path) -> list[dict]:
    import csv
    docs = []
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            if not row.get("Email") and not row.get("email"):
                continue
            # CSV path goes through the same Apollo-style mapper
            normalised = {k.strip(): v for k, v in row.items()}
            docs.append(_row_to_doc(normalised))
    return docs


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True, type=Path)
    p.add_argument("--country", default=None, help="Filter rows to one country (US/UK/IN/EU).")
    p.add_argument("--email-status", default=None,
                   help='Filter by Apollo email status (e.g., "Verified").')
    p.add_argument("--replace", action="store_true",
                   help="DROP the collection before inserting.")
    p.add_argument("--upsert", action="store_true",
                   help="Insert new + update existing by email key.")
    p.add_argument("--limit", type=int, default=None)
    args = p.parse_args()

    if not args.input.exists():
        raise SystemExit(f"File not found: {args.input}")

    print(f"Loading {args.input.name} …")
    docs = (
        load_xlsx(args.input)
        if args.input.suffix.lower() == ".xlsx"
        else load_csv(args.input)
    )
    print(f"  parsed {len(docs):,} rows with an email address")

    # Filters
    if args.country:
        before = len(docs)
        docs = [d for d in docs if d["country"] == args.country.upper()]
        print(f"  country={args.country}: {before:,} → {len(docs):,}")
    if args.email_status:
        before = len(docs)
        docs = [d for d in docs if d.get("email_status") == args.email_status]
        print(f"  email_status={args.email_status}: {before:,} → {len(docs):,}")
    if args.limit:
        docs = docs[: args.limit]
        print(f"  --limit applied: {len(docs):,}")

    # Connect — separate Mongo URI / DB from the application
    from pymongo import MongoClient, UpdateOne

    client = MongoClient(MARKETING_DB_URL, serverSelectionTimeoutMS=8000)
    db = client[MARKETING_DB_NAME]
    coll = db[COLLECTION]
    print(f"Mongo: {MARKETING_DB_URL}  →  db={MARKETING_DB_NAME}  coll={COLLECTION}")

    coll.create_index("contact_email", unique=True, background=True)
    coll.create_index("country", background=True)
    coll.create_index("status", background=True)

    if args.replace:
        deleted = coll.delete_many({}).deleted_count
        print(f"  REPLACE: dropped {deleted:,} existing rows")
        # Dedupe by email before insert_many (some Apollo exports have dupes)
        seen, deduped = set(), []
        for d in docs:
            if d["contact_email"] in seen:
                continue
            seen.add(d["contact_email"])
            deduped.append(d)
        if len(deduped) != len(docs):
            print(f"  deduped on email: {len(docs):,} → {len(deduped):,}")
        if deduped:
            res = coll.insert_many(deduped, ordered=False)
            print(f"  inserted {len(res.inserted_ids):,} rows")
    elif args.upsert:
        ops = [
            UpdateOne({"contact_email": d["contact_email"]}, {"$set": d}, upsert=True)
            for d in docs
        ]
        if ops:
            res = coll.bulk_write(ops, ordered=False)
            print(f"  upserted={res.upserted_count:,} modified={res.modified_count:,}")
    else:
        # Default: insert only new, skip duplicates
        new = 0
        for d in docs:
            try:
                coll.insert_one(d)
                new += 1
            except Exception:
                pass
        print(f"  inserted (new only): {new:,}")

    print(f"\nCurrent collection size: {coll.count_documents({}):,}")
    by_country = list(coll.aggregate([
        {"$group": {"_id": "$country", "n": {"$sum": 1}}},
        {"$sort": {"n": -1}},
    ]))
    print("By country:")
    for c in by_country:
        print(f"  {c['_id'] or '(none)':<6} {c['n']:>7,}")
    by_status = list(coll.aggregate([
        {"$group": {"_id": "$status", "n": {"$sum": 1}}},
    ]))
    print("By status:")
    for s in by_status:
        print(f"  {s['_id']:<14} {s['n']:>7,}")


if __name__ == "__main__":
    main()
